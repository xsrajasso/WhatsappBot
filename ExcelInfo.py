from openpyxl import load_workbook

def extraer_celda(ruta_archivo, nombre_hoja, celda_referencia):
    """
    Extrae el valor de una celda específica en una hoja de un archivo Excel.

    :param ruta_archivo: Ruta al archivo Excel.
    :param nombre_hoja: Nombre de la hoja de la cual extraer la celda.
    :param celda_referencia: Referencia de la celda (e.g., "B2").
    :return: Valor de la celda o None si ocurre un error.
    """
    try:
        # Cargar el libro de Excel
        wb = load_workbook(filename=ruta_archivo, data_only=True)
    except FileNotFoundError:
        print(f"Error: El archivo '{ruta_archivo}' no se encontró.")
        return None
    except Exception as e:
        print(f"Error al cargar el archivo: {e}")
        return None

    # Verificar si la hoja existe
    if nombre_hoja not in wb.sheetnames:
        print(f"Error: La hoja '{nombre_hoja}' no existe en el archivo.")
        return None

    hoja = wb[nombre_hoja]

    # Verificar si la celda existe
    try:
        celda = hoja[celda_referencia]
        valor = celda.value
        return valor
    except Exception as e:
        print(f"Error al acceder a la celda '{celda_referencia}': {e}")
        return None

# Uso del función
ruta = "Directorio.xlsx"
hoja = "1A"
celda = "I25"

valor_celda = extraer_celda(ruta, hoja, celda)
if valor_celda is not None:
    print(f"El valor de la celda {celda} en la hoja '{hoja}' es: {valor_celda}")
