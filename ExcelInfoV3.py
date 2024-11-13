from openpyxl import load_workbook

def buscar_valor_columna_a_openpyxl(ruta_archivo, nombre_hoja, valor_busqueda):
    """
    Busca un valor en la columna A de una hoja de Excel y devuelve el número de fila.

    :param ruta_archivo: Ruta al archivo Excel.
    :param nombre_hoja: Nombre de la hoja donde buscar.
    :param valor_busqueda: Valor a buscar en la columna A.
    :return: Número de fila si se encuentra, o None si no se encuentra.
    """
    try:
        # Cargar el libro de Excel
        wb = load_workbook(filename=ruta_archivo, data_only=True)
    except FileNotFoundError:
        print(f"Error: El archivo '{ruta_archivo}' no se encontró.")
        return None
    except Exception as e:
        print(f"Error al cargar el archivo: {e}")
        return None

    # Verificar si la hoja existe
    if nombre_hoja not in wb.sheetnames:
        print(f"Error: La hoja '{nombre_hoja}' no existe en el archivo.")
        return None

    hoja = wb[nombre_hoja]

    # Iterar sobre las filas de la columna A
    for fila in hoja.iter_rows(min_col=1, max_col=1):
        celda = fila[0]
        if celda.value == valor_busqueda:
            numero_fila = celda.row
            print(f"El valor '{valor_busqueda}' se encuentra en la fila {numero_fila} de Excel.")
            return numero_fila

    # Si no se encuentra el valor
    print(f"El valor '{valor_busqueda}' no se encontró en la columna A.")
    return None

def obtener_valor_celda_I_mas_fila(ruta_archivo, nombre_hoja, fila_encontrada):
    """
    Obtiene el valor de la celda en la columna I y la fila obtenida previamente.

    :param ruta_archivo: Ruta al archivo Excel.
    :param nombre_hoja: Nombre de la hoja donde buscar.
    :param fila_encontrada: Número de fila donde se encontró el valor.
    :return: Valor de la celda I + fila encontrada.
    """
    # Cargar el libro de Excel
    wb = load_workbook(filename=ruta_archivo, data_only=True)

    # Seleccionar la hoja
    hoja = wb[nombre_hoja]

    # Formar la referencia de la celda en la columna I (por ejemplo, "I5" si fila_encontrada es 5)
    referencia_celda = f"I{fila_encontrada}"

    # Obtener el valor de esa celda
    valor_celda = hoja[referencia_celda].value
    print(f"El valor de la celda {referencia_celda} es: {valor_celda}")

    return valor_celda

# Uso de la función
if __name__ == "__main__":
    ruta = "Directorio.xlsx"
    # hoja = "1A"
    # valor = 13386
    print("Escribe el grupo que quieres buscar (Ejemplo: 1A)")
    hoja = input()
    print("Escribe la matricula que quieres buscar")
    valor = int(input())

    # Buscar el valor en la columna A y obtener el número de fila
    fila_encontrada = buscar_valor_columna_a_openpyxl(ruta, hoja, valor)

    if fila_encontrada:
        # Obtener el valor de la celda I + fila_encontrada
        valor_celda_I = obtener_valor_celda_I_mas_fila(ruta, hoja, fila_encontrada)
