from openpyxl import load_workbook

def buscar_valor_en_hojas(ruta_archivo, valor_busqueda):
    """
    Busca un valor en la columna A en todas las hojas de un archivo Excel y devuelve la hoja y fila donde se encuentra.

    :param ruta_archivo: Ruta al archivo Excel.
    :param valor_busqueda: Valor a buscar en la columna A.
    :return: (nombre_hoja, número_fila) si se encuentra, o (None, None) si no se encuentra.
    """
    try:
        # Cargar el libro de Excel
        wb = load_workbook(filename=ruta_archivo, data_only=True)
    except FileNotFoundError:
        print(f"Error: El archivo '{ruta_archivo}' no se encontró.")
        return None, None
    except Exception as e:
        print(f"Error al cargar el archivo: {e}")
        return None, None

    # Iterar sobre todas las hojas del libro
    for nombre_hoja in wb.sheetnames:
        hoja = wb[nombre_hoja]

        # Iterar sobre las filas de la columna A
        for fila in hoja.iter_rows(min_col=1, max_col=1):
            celda = fila[0]
            if celda.value == valor_busqueda:
                numero_fila = celda.row
                print(f"La matricula '{valor_busqueda}' se encuentra en la fila {numero_fila} del grupo '{nombre_hoja}'.")
                return nombre_hoja, numero_fila

    # Si no se encuentra el valor
    print(f"El valor '{valor_busqueda}' no se encontró en ninguna hoja del directorio.")
    return None, None

def obtener_valor_celda_B_y_I(ruta_archivo, nombre_hoja, fila_encontrada):
    """
    Obtiene los valores de las celdas en las columnas B e I en la fila obtenida previamente.

    :param ruta_archivo: Ruta al archivo Excel.
    :param nombre_hoja: Nombre de la hoja donde buscar.
    :param fila_encontrada: Número de fila donde se encontró el valor.
    :return: Valores de las celdas B y I + fila encontrada.
    """
    # Cargar el libro de Excel
    wb = load_workbook(filename=ruta_archivo, data_only=True)

    # Seleccionar la hoja
    hoja = wb[nombre_hoja]

    # Obtener el valor de la celda en la columna B (nombre del alumno)
    referencia_celda_B = f"B{fila_encontrada}"
    valor_celda_B = hoja[referencia_celda_B].value

    # Obtener el valor de la celda en la columna I (telefono de la mama)
    referencia_celda_I = f"I{fila_encontrada}"
    valor_celda_I = hoja[referencia_celda_I].value

    # Obtener el valor de la celda en la columna I (telefono del papa)
    referencia_celda_F = f"F{fila_encontrada}"
    valor_celda_F = hoja[referencia_celda_F].value

    print(f"El alumno en la celda {referencia_celda_B} es: {valor_celda_B}")
    print(f"El telefono de la mama es : {valor_celda_I}")
    print(f"El telefono del papa es : {valor_celda_F}")

    return valor_celda_B, valor_celda_I, valor_celda_F

def escribir_mensaje_en_archivo(nombre_alumno, telefonoMama, telefonoPapa, matricula, ruta_archivo_salida):
    """
    Escribe un mensaje en un archivo de texto con la información del alumno.

    :param nombre_alumno: Nombre del alumno.
    :param matricula: Matrícula del alumno.
    :param ruta_archivo_salida: Ruta al archivo de salida donde se escribirá el mensaje.
    """
    mensaje = (
        f"Enviar al telfono de la mamá: ({telefonoMama}). \n"
        f"Enviar al telfono del papá: ({telefonoPapa}). \n"
        "\n"
        f"Para notificarle que el día de hoy el alumno: {nombre_alumno} con matrícula {matricula}, "
        "tuvo más de tres inasistencias a clases, lo cual pone en riesgo su formación académica. "
        "Para más información por favor comuníquese con la Lic. Liliana Ávila a los teléfonos "
        "8443275869 vía WhatsApp o bien al 8447771002 para llamadas.\nGracias."
    )
    
    with open(ruta_archivo_salida, 'w') as archivo_salida:
        archivo_salida.write(mensaje)
    
    print(f"Mensaje escrito en el archivo {ruta_archivo_salida}.")
    # input("Type to exit...")

# Uso de las funciones
if __name__ == "__main__":
    ruta = "Directorio.xlsx"
    # hoja = "1A"
    # valor = 13386
    # print("Escribe el grupo que quieres buscar (Ejemplo: 1A)")
    # hoja = input()
    print("Escribe la matricula que quieres buscar")
    valor = int(input())
    print("Buscando...")
    ruta_salida = "mensajeSalida.txt"

    # Buscar el valor en todas las hojas y obtener el nombre de la hoja y el número de fila
    nombre_hoja_encontrada, fila_encontrada = buscar_valor_en_hojas(ruta, valor)

    if nombre_hoja_encontrada and fila_encontrada:
        # Obtener los valores de las celdas B e I + fila_encontrada
        nombre_alumno, telefonoMama, telefonoPapa = obtener_valor_celda_B_y_I(ruta, nombre_hoja_encontrada, fila_encontrada)

        # Escribir el mensaje en el archivo de salida
        escribir_mensaje_en_archivo(nombre_alumno, telefonoMama, telefonoPapa, valor, ruta_salida)
