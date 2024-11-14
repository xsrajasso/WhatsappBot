from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
import sys

# Obtener la fecha actual y el día de la semana
fecha_actual = datetime.now()
dia_semana_numero = fecha_actual.weekday()
dias_semana = ["Lunes", "Martes", "Miercoles", "Jueves", "Viernes", "Sabado", "Domingo"]
dia_semana_nombre = dias_semana[dia_semana_numero]
# dia_semana_nombre = "Sabado"

# Función para obtener matrículas con más de 3 faltas como enteros
def obtener_matriculas_con_faltas(archivo, hoja, umbral_faltas=2):
    df = pd.read_excel(archivo, sheet_name=hoja)
    # print("Hoy es:", dia_semana_nombre)
    matriculas = []

    if dia_semana_nombre == "Lunes":
        matriculas = df.loc[df['Faltas'] > umbral_faltas, 'Matricula'].astype(int).tolist()
    elif dia_semana_nombre == "Martes":
        matriculas = df.loc[df['Faltas.1'] > umbral_faltas, 'Matricula'].astype(int).tolist()
    elif dia_semana_nombre == "Miercoles":
        matriculas = df.loc[df['Faltas.2'] > umbral_faltas, 'Matricula'].astype(int).tolist()
    elif dia_semana_nombre == "Jueves":
        matriculas = df.loc[df['Faltas.3'] > umbral_faltas, 'Matricula'].astype(int).tolist()
    elif dia_semana_nombre == "Viernes":
        matriculas = df.loc[df['Faltas.4'] > umbral_faltas, 'Matricula'].astype(int).tolist()
    else:
        print("Hoy no hay clases! Presione enter para salir.")
        input()
        sys.exit()
        
    return matriculas

def buscar_valor_en_hojas(ruta_archivo, valor_busqueda):
    try:
        wb = load_workbook(filename=ruta_archivo, data_only=True)
    except FileNotFoundError:
        print(f"Error: El archivo '{ruta_archivo}' no se encontró.")
        return None, None
    except Exception as e:
        print(f"Error al cargar el archivo: {e}")
        return None, None

    for nombre_hoja in wb.sheetnames:
        hoja = wb[nombre_hoja]
        for fila in hoja.iter_rows(min_col=1, max_col=1):
            celda = fila[0]
            if celda.value == valor_busqueda:
                numero_fila = celda.row
                print(f"La matricula '{valor_busqueda}' se encuentra en la fila {numero_fila} del grupo '{nombre_hoja}'.")
                return nombre_hoja, numero_fila

    print(f"La matricula '{valor_busqueda}' no se encontró en ninguna hoja del directorio.")
    return None, None

def obtener_valor_celda_B_y_I(ruta_archivo, nombre_hoja, fila_encontrada):
    wb = load_workbook(filename=ruta_archivo, data_only=True)
    hoja = wb[nombre_hoja]
    valor_celda_B = hoja[f"B{fila_encontrada}"].value
    valor_celda_I = hoja[f"I{fila_encontrada}"].value
    valor_celda_F = hoja[f"F{fila_encontrada}"].value

    print(f"El alumno en la celda B{fila_encontrada} es: {valor_celda_B}")
    print(f"El telefono de la mama es : {valor_celda_I}")
    print(f"El telefono del papa es : {valor_celda_F}")

    return valor_celda_B, valor_celda_I, valor_celda_F

def escribir_mensaje_en_archivo(nombre_alumno, telefonoMama, telefonoPapa, matricula, ruta_archivo_salida):
    mensaje = (
        f"Enviar al teléfono de la mamá: ({telefonoMama}).\n"
        f"Enviar al teléfono del papá: ({telefonoPapa}).\n\n"
        f"¡Buen dia! Nos comunicamos de parte de Prefectura de bachillerato Carolina. " 
        "Para notificarle que el día de hoy el alumno: {nombre_alumno} con matrícula {matricula}, "
        "tuvo más de tres inasistencias a clases, lo cual pone en riesgo su formación académica. "
        "Para más información por favor comuníquese con la Lic. Liliana Ávila a los teléfonos "
        "8443275869 vía WhatsApp o bien al 8447771002 para llamadas.\n\nGracias.\n\n"
        "-----------------------------------------------------------------------------------------\n\n"
    )
    
    # Agregar el mensaje al archivo en modo 'append'
    with open(ruta_archivo_salida, 'a') as archivo_salida:
        archivo_salida.write(mensaje)
    
    print(f"Mensaje agregado al archivo {ruta_archivo_salida}.")

def escribir_mensaje_no_encontrado(matricula, ruta_archivo_salida):
    mensaje = f"La matrícula {matricula} no se encontró en el directorio.\n\n"
    with open(ruta_archivo_salida, 'a') as archivo_salida:
        archivo_salida.write(mensaje)
    print(f"Mensaje de no encontrado para la matrícula {matricula} agregado al archivo {ruta_archivo_salida}.")

# Especifica el archivo y la hoja
archivo = "PEGGY C17-C24.xlsx"
print("¿Qué grupo quieres buscar? (Ejemplo: 1°A)")
hoja = input()
matriculas = obtener_matriculas_con_faltas(archivo, hoja)
ruta_salida = "mensajeSalida.txt"

# Vaciar el archivo de salida antes de agregar nuevos mensajes
open(ruta_salida, 'w').close()

# Para cada matrícula, busca y genera el mensaje
if not matriculas:  # Si la lista está vacía
    print(f"No hay alumnos con al menos 3 faltas en el grupo {hoja} el día de hoy {dia_semana_nombre}.")
else:
    print(f"Se encontraron alumnos con al menos 3 faltas en el día de hoy {dia_semana_nombre}.")
    for m in matriculas:
        print(f"Se está buscando la matrícula ({m}) en el Directorio...")
        ruta_directorio = "Directorio.xlsx"
        nombre_hoja_encontrada, fila_encontrada = buscar_valor_en_hojas(ruta_directorio, m)
    
        if nombre_hoja_encontrada and fila_encontrada:
            nombre_alumno, telefonoMama, telefonoPapa = obtener_valor_celda_B_y_I(ruta_directorio, nombre_hoja_encontrada, fila_encontrada)
            escribir_mensaje_en_archivo(nombre_alumno, telefonoMama, telefonoPapa, m, ruta_salida)
        else:
            escribir_mensaje_no_encontrado(m, ruta_salida)
    
    print(f"Mensajes completados y guardados en {ruta_salida}.")
